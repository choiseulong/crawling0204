# 페이지가 역순으로 시작해서 10페이지가 넘는 url이 처음에 바로 선택할 수 없음을 막바지에 알아서 수정을 하지 못했던게 아쉬워요.
# 처음에는 1페이지만 따로 짜고 나머지 2페이지부터 해서 문제없었는데 하나로 합치는 과정에서 생각지못했네요.
# 또 grid를 사용했는데 중앙정렬까지 못구현했어요.
# 또 엑셀로 저장시 한글에 링크를 다는 법까지 못찾았어요.
# 또 몇개의 기사가 검색되었는지 출력하려고 했는데 이 역시 막바지에 못했어요.
# 또 구글 기사 이미지가 100px 100px 여서 조금 키워 가져왔는데 좌우 여백이 하얀색이라 검정으로 바꾸고 싶었는데 실패했어요.
# 엔터키 눌렀을 때 input 내용이 ajax를 통해서 함수로 가게 하고 싶었는데 자바스크립트로된 내용은 봤지만 jquery에서 본 내용으로만으로
# 짜는데 실패했어요.
# 검색내역 + 페이지네이션 까지 했으면 좋았을탠데 아쉬웠어요.
# 매번 엑셀로 저장할 때 다른 파일로 저장되게 하지 못해 아쉬웠어요.
# alert를 꾸며보고 싶었는데 역시 막바지에 생각나서 못해봤어요.


from selenium import webdriver
from bs4 import BeautifulSoup
from flask import Flask, render_template, jsonify, request
from openpyxl import Workbook, load_workbook

app = Flask(__name__, static_url_path='/static')


@app.route('/search', methods=["POST"])
def search():
    google_list = []
    google_list_href = []
    google_list_img = []
    more_info = []
    search_number = request.form['search_number']
    SN = int(search_number) + 1
    nums = int(str(SN-1)+"0")
    if request.method == "POST":

        driver = webdriver.Chrome('./chromedriver')
        driver.implicitly_wait(3)
        driver.get('https://google.com')
        driver.find_element_by_name("q").send_keys(request.form['search_name'])

        # input 에서 받은 'search_name' 으로 검색

        driver.find_element_by_name("btnK").click()
        soup = BeautifulSoup(driver.page_source, 'html.parser')

        more_info_all = soup.find_all("a", {'class': 'f9UGee q qs'})
        for name in more_info_all:
            more_info.append(name.text)
        for num in range(len(more_info)):
            if more_info[num] == "뉴스":
                num = num + 1
                driver.find_element_by_xpath("//*[@id='ow15']/a").click()
                driver.find_element_by_xpath("//*[@id='lb']/div/a[" + str(num) + "]").click()
            else:
                pass
        # 더보기에 뉴스가 있다면 그걸 먼저 클릭.
        # (뉴스가 더보기 속에 숨어 있는 경우에요! 위치가 키워드마다 달라서 하나씩 훝었어요.)

        for i in range(2, 6):
            click_check = soup.select("#hdtb-msb-vis > div:nth-child(" + str(i) + ")")
            for j in click_check:
                if j.text == "뉴스":
                    driver.find_element_by_xpath("//*[@id='hdtb-msb-vis']/div[" + str(i) + "]/a").click()
                else :
                    pass
        # 더보기에 뉴스가 없다면 항목의 네 가지 위치에서 검색.
        # (역시나 뉴스 위치가 검색어에 따라 달라져서 하나하나 훝어야 됐어요.)

        for i in range(SN, 1, -1) :
            driver.find_element_by_xpath("//*[@id='nav']/tbody/tr/td[" + str(i) + "]/a").click()
            soup = BeautifulSoup(driver.page_source, 'html.parser')
        # 첫 페이지는 .click이 활성화 되지 않아 긁어오고 싶은 마지막 페이지에서 역순으로 첫 번째 페이지 까지 크롤링 했어요.
            for names in range(1, 11):
                list_name = soup.select("#rso > div > div:nth-child(" + str(names) + ") > div > div.gG0TJc > h3 > a")
                for listName in list_name:
                    google_list.append(listName.text.strip())
            # 뉴스 제목 긁어오기
            for links in soup.find_all("a", {'class': 'l lLrAF'}):
                if 'href' in links.attrs:
                    google_list_href.append(links.attrs['href'])
            # 뉴스 링크
            for img in soup.find_all("img", {'class': 'th BbeB2d'}):
                if 'src' in img.attrs:
                    google_list_img.append(img.attrs['src'])
            # 뉴스 대표이미지

        write_wb = Workbook()
        write_ws = write_wb.active
        for i in range(len(google_list)):
            write_ws.cell(i+1, 1, google_list[i])
            write_ws.cell(i+1, 2, google_list_href[i])
        write_wb.save("google_list.xlsx")
            # 엑셀로 저장 링크는 어찌걸지..?

    return jsonify(
        {'status': 'OK', 'list': google_list, 'list_href': google_list_href, 'list_img_src': google_list_img, 'nums': nums})


@app.route('/')
def index():
    return render_template('index.html')


if __name__ == '__main__':
    app.run()
